import os
import logging
import re
import openpyxl
from enums import CMSPathTypes, CMSSubmissionsFileExcelColumns, CMSTools
from path_finder import PathFinder
from map_path_builder import MapPathBuilder
import shutil
from constants import DRIVE_LETTER, CMS_FOLDER
from rich.console import Console
from rich.prompt import Prompt


def handle_cms_path_builder(submissions_file_path: str, path_type: str) -> str:
    """
   Handle building CMS (Content Management System) paths based on the provided submissions file and path type.

   :param submissions_file_path: The path to the Excel file containing submission information.
   :param path_type: The type of CMS path to build.

   :return: A string indicating the outcome of building CMS paths.
       Possible return values:
       - "success": CMS path building was successful.
       - "error - file path": The provided file path is invalid.
       - "error - invalid cms path type": The provided CMS path type is invalid.
       - "error - excel file columns": The columns in the Excel file are invalid.
       - "error - cms path": The CMS path is invalid.
   """

    if not os.path.isfile(submissions_file_path):
        return "error - file path"

    if not submissions_file_path.endswith(".xlsx"):
        return "error - file path"

    if not os.path.exists("Y:\\HC"):
        return "error - not connected to vpn and oes"

    cms_path = os.path.join(DRIVE_LETTER, CMS_FOLDER)
    if not os.path.isdir(cms_path):
        return "error - cms path"

    valid_path_type = False

    for item in CMSPathTypes:
        if item.value == path_type:
            valid_path_type = True
            break

    if not valid_path_type:
        return "error - invalid cms path type"

    wb = openpyxl.load_workbook(submissions_file_path)
    ws = wb.active

    path_builder = MapPathBuilder()
    path_finder = PathFinder()

    submissions_col = ws.cell(row=1, column=1).value
    source_col = ws.cell(row=1, column=2).value

    if submissions_col is None or source_col is None:
        return "error - excel file columns"

    if submissions_col.lower() != CMSSubmissionsFileExcelColumns.SUBMISSION.value.lower() or source_col.lower() != CMSSubmissionsFileExcelColumns.SOURCE.value.lower():
        return "error - excel file columns"

    ws.cell(row=1, column=3).value = CMSSubmissionsFileExcelColumns.DESTINATION.value

    row_count = 2
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if path_type == CMSPathTypes.PRODUCT.value:
            matches = re.findall(r"\b\d{6}", str(row[0]).lower())
            # File and submission number
            if len(matches) == 2:
                try:
                    path = path_finder.find_product_folder(path_builder.build_product_path(str(matches[0])),
                                                           str(matches[0]),
                                                           str(matches[1]))
                    ws.cell(row=row_count, column=3).value = path if path is not None else ""

                except FileNotFoundError:
                    ws.cell(row=row_count, column=3).value = ""

            # File number only
            elif len(matches) == 1:
                try:
                    path = path_finder.find_product_folder(path_builder.build_product_path(str(matches[0])),
                                                           str(matches[0]))
                    ws.cell(row=row_count, column=3).value = path if path is not None else ""

                except FileNotFoundError:
                    ws.cell(row=row_count, column=3).value = ""

        elif path_type == CMSPathTypes.PRODUCT_POST_LICENCE_FOLDER.value:
            matches = re.findall(r"\b\d{6}", str(row[0]).lower())
            try:
                # Look for post licence folder
                path = path_finder.find_product_post_licence_folder(
                    path_builder.build_product_path(str(matches[0])),
                    str(matches[0]))
                ws.cell(row=row_count, column=3).value = path if path is not None else ""

            except FileNotFoundError:
                ws.cell(row=row_count, column=3).value = ""

        elif path_type == "Site" or path_type == "Foreign Site":
            return "error"
        elif path_type == "Trading Partner":
            return "error"
        elif path_type == "Clinical Trial":
            return "error"
        elif path_type == "Company":
            return "error"
        elif path_type == "Master File":
            return "error"

        row_count += 1

    wb.save(submissions_file_path)
    return "success"


def handle_bulk_uploader(file_path: str, generate_log_file: bool, create_missing_paths: bool) -> str:
    """
    Handle bulk uploading of files to the CMS based on the information provided in the Excel file.

    :param file_path: The path to the Excel file containing file upload information.
    :param generate_log_file: A boolean indicating whether to generate a log file during the upload process.
    :param create_missing_paths: A boolean indicating whether to create missing paths in the CMS during the upload process.

    :return: A string indicating the outcome of the bulk uploading process.
        Possible return values:
        - "success": Bulk uploading was successful.
        - "error - file path": The provided file path is invalid.
        - "error - excel file columns": The columns in the Excel file are invalid.
        - "error - cms path": The CMS path is invalid.
    """
    if not os.path.isfile(file_path):
        return "error - file path"

    if not file_path.endswith(".xlsx"):
        return "error - file path"

    if not os.path.exists("Y:\\HC"):
        return "error - not connected to vpn and oes"

    cms_path = os.path.join(DRIVE_LETTER, CMS_FOLDER)
    if not os.path.isdir(cms_path):
        return "error - cms path"

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    submissions_col = ws.cell(row=1, column=1).value
    source_col = ws.cell(row=1, column=2).value
    dest_col = ws.cell(row=1, column=3).value

    if submissions_col is None or source_col is None or dest_col is None:
        return "error - excel file columns"

    if (submissions_col.lower() != CMSSubmissionsFileExcelColumns.SUBMISSION.value.lower() or
            source_col.lower() != CMSSubmissionsFileExcelColumns.SOURCE.value.lower() or
            dest_col.lower() != CMSSubmissionsFileExcelColumns.DESTINATION.value.lower()):
        return "error - excel file columns"

    # Configure the logger
    if generate_log_file:
        logging.basicConfig(level=logging.INFO, filename="bulkUploader.log", filemode="w",
                            format="%(asctime)s - %(levelname)s - %(message)s")
        logging.info("File upload started...")

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        try:
            print(f"{str(row[0])}")
            print(f"Working on copying {row[1]} to {row[2]}")
            if generate_log_file:
                logging.info(f"Working on copying {row[1]} to {row[2]}")

            source_file = os.path.basename(row[1])
            # Check if row destination is empty
            if row[2] is None:
                print("Row destination is empty! Skipping...")
                if generate_log_file:
                    logging.info("Row destination is empty! Skipping...")

            # Check CMS to see if the folder path of the destination exists
            elif not os.path.exists(row[2]):
                print("Destination folder path doesn't exist in CMS!")
                if generate_log_file:
                    logging.info("Destination folder path doesn't exist in CMS!")

                # Create the path
                if create_missing_paths:
                    # Create the necessary directories (destination)
                    os.makedirs(row[2])
                    # Copy source file to the destination
                    shutil.copy2(row[1], row[2])
                    print("Created missing path and copied file to it!")
                    if generate_log_file:
                        logging.info("Created missing path and copied file to it!")

                # If the folder create missing paths is not checked exist skip
                else:
                    print("Skipping...")
                    if generate_log_file:
                        logging.info("Skipping...")

            # Check CMS to see if the source file already exists at the destination
            elif not os.path.exists(os.path.join(row[2], source_file)):
                # Copy source file to the destination
                shutil.copy2(row[1], row[2])
                print("File copied successfully!")
                if generate_log_file:
                    logging.info("File copied successfully!")

            # If the file exists don't overwrite it
            else:
                print("File already exists at the destination! No creation necessary...")
                if generate_log_file:
                    logging.info("File already exists at the destination! No creation necessary...")
        except Exception as e:
            # Log the error
            if generate_log_file:
                logging.error(str(e))

    return "success"


def run_app():
    console = Console()
    console.print("CMS Automation Tool", style="bold green")
    console.print(f"{'-'*50}", style="bold blue")

    while True:

        tool_selection_input = Prompt.ask(prompt="[bold green]Which tool would you like to use?[/bold green]", choices=["Path Builder", "Bulk Uploader"], show_choices=True, case_sensitive=False, console=console)

        results = ""

        if tool_selection_input.lower() == CMSTools.PATH_BUILDER.value.lower():
            file_path_input = Prompt.ask("[bold green]Enter path to the file containing submissions[/bold green]", console=console)
            type_choices = CMSPathTypes.get_values()
            path_type_input = Prompt.ask(prompt="[bold green]Enter the type of path to build", choices=type_choices, show_choices=True, case_sensitive=False, console=console)
            results = handle_cms_path_builder(file_path_input, path_type_input)

        elif tool_selection_input.lower() == CMSTools.BULK_UPLOADER.value.lower():
            file_path_input = Prompt.ask("[bold green]Enter path to the file containing submissions, along with their source and destination information[/bold green]", console=console)
            results = handle_bulk_uploader(file_path_input, True, True)

        console.print(f"{results}", style="bold green")
        run_another_input = Prompt.ask(prompt="[bold blue]Would you like to run another tool?[/bold blue]", choices=["Yes", "No"], show_choices=True, case_sensitive=False, console=console)

        if run_another_input.lower() == "no":
            console.print(f"Exiting...", style="bold red")
            break

